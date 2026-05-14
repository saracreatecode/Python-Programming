"""
SCRIPT 1: LIMPIEZA Y VALIDACIÓN DEL EXCEL DE CONTACTOS
=======================================================
Autor: generado para tu empresa
Versión: 1.0

Qué hace:
- Lee todas las hojas del Excel (A-H, excepto G)
- Normaliza y limpia los datos (teléfonos, emails, textos)
- Detecta duplicados por email entre hojas
- Añade columna ACTIVO si no existe
- Genera un log detallado de errores y avisos
- Guarda un Excel limpio listo para sincronizar

Requisitos:
    pip install openpyxl pandas

Uso:
    python 01_limpiar_excel.py
"""

import pandas as pd
import openpyxl
import re
import os
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIGURACIÓN — ajusta estas rutas
# ─────────────────────────────────────────────

RUTA_EXCEL_ORIGINAL = r"\\servidor\ruta\al\excel\original.xlsx"   # Ruta al Excel en el servidor
RUTA_EXCEL_LIMPIO   = r"\\servidor\ruta\al\excel\limpio.xlsx"  # Dónde guardar el resultado
RUTA_LOG            = r"\\servidor\ruta\al\log\limpieza_log.txt"  # Log de errores/avisos

# Hojas a procesar (G se excluye intencionadamente)
HOJAS = {
    "A": {"cabecera_fila": 4, "nombre": "Clientes_ActoresClave"},
    "B": {"cabecera_fila": 3, "nombre": "Proveedores"},
    "C": {"cabecera_fila": 3, "nombre": "Ingenieros_Consultores_Legal"},
    "D": {"cabecera_fila": 3, "nombre": "Constructoras"},
    "E": {"cabecera_fila": 3, "nombre": "Arquitectos_Municipales"},
    "F": {"cabecera_fila": 3, "nombre": "Oficina_Consumibles"},
    "H": {"cabecera_fila": 3, "nombre": "Internacionales"},
}

# Mapeo de columnas originales → nombres internos normalizados
# Ajusta si alguna columna tiene nombre ligeramente diferente en tu Excel
MAPEO_COLUMNAS = {
    "NOMBRE":               "nombre",
    "1er APELLIDO":         "apellido1",
    "2º APELLIDO":          "apellido2",
    "EMPRESA / RAZÓN SOCIAL": "empresa",
    "TIPO":                 "tipo",
    "ÁREA":                 "area",
    "CARGO":                "cargo",
    "MAIL":                 "MAIL1",
    "MAIL1":                "MAIL1",
    "MAIL2":                "MAIL2",
    "MAIL3":                "MAIL3",
    "AÑO ÚLTIMO TRABAJO":   "año_ultimo_trabajo",
    "NOTA":                 "nota",
    "DIRECCIÓN":            "direccion",
    "WEB":                  "web",
    "ACCIONES 1":           "accion1",
    "ACCIONES 2":           "accion2",
    "ACCIONES 3":           "accion3",
    "ACCIONES 4":           "accion4",
    "ACCIONES 5":           "accion5",
    "ACCIONES 6":           "accion6",   # Solo hoja A
    "PAÍS":                 "pais",      # Solo hoja H
}

# ─────────────────────────────────────────────
# FUNCIONES DE LIMPIEZA
# ─────────────────────────────────────────────

# Caracteres invisibles Unicode que pueden colarse al copiar/pegar
CHARS_INVISIBLES = [
    '\u200e',  # Left-to-Right Mark (LRM)
    '\u200f',  # Right-to-Left Mark (RLM)
    '\u200b',  # Zero Width Space
    '\u200c',  # Zero Width Non-Joiner
    '\u200d',  # Zero Width Joiner
    '\u00ad',  # Soft Hyphen
    '\ufeff',  # BOM / Zero Width No-Break Space
    '\u202a',  # Left-to-Right Embedding
    '\u202b',  # Right-to-Left Embedding
    '\u202c',  # Pop Directional Formatting
]

def limpiar_texto(valor):
    """Elimina espacios extra, caracteres invisibles Unicode y normaliza."""
    if pd.isna(valor) or str(valor).strip() == "":
        return ""
    texto = str(valor).strip()
    # Eliminar caracteres invisibles
    for char in CHARS_INVISIBLES:
        texto = texto.replace(char, "")
    texto = texto.strip()
    # Elimina múltiples espacios internos
    texto = re.sub(r'\s+', ' ', texto)
    return texto

def limpiar_nombre_propio(valor):
    """Normaliza nombres propios: primera letra mayúscula en cada palabra."""
    texto = limpiar_texto(valor)
    if not texto:
        return ""
    # Excepciones: partículas que van en minúscula (de, del, la, los...)
    particulas = {"de", "del", "la", "las", "los", "el", "y", "e", "i"}
    palabras = texto.lower().split()
    resultado = []
    for i, palabra in enumerate(palabras):
        if i == 0 or palabra not in particulas:
            resultado.append(palabra.capitalize())
        else:
            resultado.append(palabra)
    return " ".join(resultado)

def limpiar_email(valor):
    """
    Valida y normaliza email. Devuelve (email_limpio, es_valido, mensaje_error).
    Soporta:
    - Guiones en parte local y dominio: nombre-apellido@empresa-web.com
    - Caracteres internacionales en parte local: ß, ü, ñ, etc. (RFC 6531)
    - Dominios con guión: my-company.com
    """
    texto = limpiar_texto(valor).strip()
    if not texto:
        return "", None, None  # Vacío es aceptable

    # Normalizar: lowercase solo en el dominio, respetar parte local
    if "@" in texto:
        parte_local, dominio = texto.rsplit("@", 1)
        texto_normalizado = parte_local + "@" + dominio.lower()
    else:
        texto_normalizado = texto

    # Patrón extendido con soporte unicode (ß, ñ, ü...) y guiones
    patron = r'^[\w.%+\-]+@[a-zA-Z0-9\-]+(\.[a-zA-Z0-9\-]+)*\.[a-zA-Z]{2,}$'

    if re.match(patron, texto_normalizado, re.UNICODE):
        return texto_normalizado, True, None
    else:
        return texto_normalizado, False, f"Email con formato inválido: '{texto_normalizado}'"

def limpiar_telefono(valor):
    """
    Limpia un teléfono: elimina caracteres no numéricos excepto + al inicio.
    Devuelve (telefono_limpio, es_valido, mensaje_error).
    """
    texto = limpiar_texto(str(valor)) if not pd.isna(valor) else ""
    if not texto or texto == "nan":
        return "", None, None  # Vacío aceptable

    # Conserva el + inicial si existe (prefijo internacional)
    tiene_mas = texto.startswith("+")
    # Elimina todo excepto dígitos
    solo_digitos = re.sub(r'\D', '', texto)

    if not solo_digitos:
        return "", False, f"Teléfono sin dígitos válidos: '{texto}'"

    # Reconstruye con + si lo tenía
    limpio = ("+" if tiene_mas else "") + solo_digitos

    # Números cortos de servicios públicos españoles (010, 011, 012, 016, 060, 112...)
    # son válidos aunque tengan menos de 7 dígitos
    NUMEROS_CORTOS_VALIDOS = {"010", "011", "012", "016", "017", "060", "061", "062", "112", "016"}
    if solo_digitos in NUMEROS_CORTOS_VALIDOS:
        return limpio, True, None

    # Validación de longitud: mínimo 3 dígitos (otros servicios cortos), máximo 15 (ITU-T E.164)
    if len(solo_digitos) < 3 or len(solo_digitos) > 15:
        return limpio, False, f"Teléfono con longitud inusual ({len(solo_digitos)} dígitos): '{limpio}'"

    return limpio, True, None

def limpiar_web(valor):
    """Normaliza URLs: añade https:// si falta."""
    texto = limpiar_texto(valor).lower()
    if not texto:
        return ""
    if texto and not texto.startswith(("http://", "https://")):
        texto = "https://" + texto
    return texto

def limpiar_anio(valor):
    """Valida que el año sea un número de 4 dígitos razonable."""
    texto = limpiar_texto(str(valor)) if not pd.isna(valor) else ""
    if not texto or texto == "nan":
        return None, None, None  # None en lugar de "" para columnas numéricas
    try:
        anio = int(float(texto))
        if 1950 <= anio <= datetime.now().year:
            return str(anio), True, None
        else:
            return str(anio), False, f"Año fuera de rango razonable: '{anio}'"
    except (ValueError, TypeError):
        return texto, False, f"Año con valor no numérico: '{texto}'"


# ─────────────────────────────────────────────
# LEER EXCEL CON COLUMNAS TLF COMBINADAS
# ─────────────────────────────────────────────

def leer_hoja_con_tlf(ruta_excel, nombre_hoja, fila_cabecera):
    """
    Lee una hoja del Excel desde la fila de cabecera indicada.
    - Los datos empiezan en columna B (columna A ignorada).
    - Las cabeceras se limpian de caracteres invisibles.
    - Las columnas sin nombre se descartan.
    - Guarda el número de fila real del Excel en columna __fila_excel__
      para que el log muestre siempre el número correcto.
    """
    CHARS_INV = ['\u200e','\u200f','\u200b','\u200c','\u200d',
                 '\u00ad','\ufeff','\u202a','\u202b','\u202c']

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb[nombre_hoja]

    # Leer cabecera desde columna B
    cabecera_raw = [c.value for c in ws[fila_cabecera][1:]]  # [1:] = desde col B

    # Limpiar nombres de cabecera
    cabecera_limpia = []
    for col in cabecera_raw:
        if col is None:
            cabecera_limpia.append(None)
        else:
            nombre = str(col).strip()
            for char in CHARS_INV:
                nombre = nombre.replace(char, "")
            cabecera_limpia.append(nombre.strip() if nombre.strip() else None)

    # Leer filas de datos guardando el número de fila real del Excel
    registros = []
    for fila_excel_num, row in enumerate(
        ws.iter_rows(min_row=fila_cabecera + 1, min_col=2, values_only=True),
        start=fila_cabecera + 1
    ):
        valores = list(row)
        # Saltar filas completamente vacías
        if all(v is None for v in valores):
            continue
        registro = dict(zip(cabecera_limpia, valores))
        registro["__fila_excel__"] = fila_excel_num  # número real en el Excel
        registros.append(registro)

    if not registros:
        wb.close()
        return pd.DataFrame()

    df = pd.DataFrame(registros)

    # Eliminar columnas sin nombre (None)
    df = df.loc[:, df.columns.notna()]

    # Resetear índice secuencial (el número de fila real está en __fila_excel__)
    df = df.reset_index(drop=True)

    wb.close()
    return df


# ─────────────────────────────────────────────
# PROCESAMIENTO PRINCIPAL
# ─────────────────────────────────────────────

def procesar_excel():
    log_lineas = []
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_lineas.append(f"{'='*60}")
    log_lineas.append(f"LOG LIMPIEZA EXCEL — {timestamp}")
    log_lineas.append(f"{'='*60}\n")

    # Verificar que el fichero existe
    if not os.path.exists(RUTA_EXCEL_ORIGINAL):
        print(f"ERROR: No se encuentra el Excel en: {RUTA_EXCEL_ORIGINAL}")
        return

    # Cargar libro para obtener nombres reales de hojas
    wb_check = openpyxl.load_workbook(RUTA_EXCEL_ORIGINAL, read_only=True)
    hojas_disponibles = wb_check.sheetnames
    wb_check.close()
    print(f"Hojas encontradas en el Excel: {hojas_disponibles}\n")

    hojas_procesadas = {}          # nombre_interno -> DataFrame limpio
    emails_globales = {}           # email -> lista de hojas donde aparece

    # ── Procesar cada hoja ──────────────────────────────────────
    for letra, config in HOJAS.items():
        # Buscar la hoja por letra inicial o nombre parcial
        hoja_real = None
        for h in hojas_disponibles:
            if h.strip().upper().startswith(letra):
                hoja_real = h
                break

        if not hoja_real:
            log_lineas.append(f"⚠️  HOJA {letra}: No encontrada en el Excel. Se omite.\n")
            continue

        nombre_interno = config["nombre"]
        fila_cabecera  = config["cabecera_fila"]

        print(f"Procesando hoja '{hoja_real}' → {nombre_interno}...")
        log_lineas.append(f"{'─'*50}")
        log_lineas.append(f"HOJA {letra}: {hoja_real} ({nombre_interno})")
        log_lineas.append(f"{'─'*50}")

        df = leer_hoja_con_tlf(RUTA_EXCEL_ORIGINAL, hoja_real, fila_cabecera)

        if df.empty:
            log_lineas.append("  ⚠️  Hoja vacía o sin datos.\n")
            continue

        # Renombrar columnas al esquema interno normalizado
        renombres = {}
        for col_original, col_interno in MAPEO_COLUMNAS.items():
            # Búsqueda flexible (sin distinguir mayúsculas ni espacios extra)
            for col_df in df.columns:
                if col_df.strip().upper() == col_original.strip().upper():
                    renombres[col_df] = col_interno
                    break
        df = df.rename(columns=renombres)

        # Añadir columnas faltantes como vacías
        for col in ["nombre","apellido1","apellido2","empresa","tipo","area","cargo",
                    "MAIL1","MAIL2","MAIL3","TLF1","TLF2","TLF3",
                    "año_ultimo_trabajo","nota","direccion","web",
                    "accion1","accion2","accion3","accion4","accion5","accion6","pais"]:
            if col not in df.columns:
                df[col] = ""

        # Añadir columna ACTIVO si no existe
        if "ACTIVO" not in df.columns:
            df["ACTIVO"] = "SI"
        else:
            df["ACTIVO"] = df["ACTIVO"].fillna("SI").astype(str).str.upper().str.strip()

        # Añadir columna CATEGORIA con el nombre de la hoja
        df["categoria"] = nombre_interno

        # ── Limpiar fila por fila ───────────────────────────────
        errores_hoja = 0
        avisos_hoja  = 0

        for idx, fila in df.iterrows():
            # Número de fila real en el Excel, guardado durante la lectura
            fila_num = int(fila.get("__fila_excel__", idx + fila_cabecera + 1))
            avisos_fila = []

            # Nombre — si está vacío, usar empresa silenciosamente sin aviso en log
            nombre_limpio = limpiar_nombre_propio(fila.get("nombre", ""))
            empresa_limpia = limpiar_texto(fila.get("empresa", ""))
            if not nombre_limpio:
                if empresa_limpia:
                    nombre_limpio = empresa_limpia
                    # Sin aviso en log: comportamiento esperado para contactos de empresa
                else:
                    log_lineas.append(f"  ❌ FILA {fila_num}: Nombre y empresa vacíos → contacto OMITIDO")
                    df.at[idx, "ACTIVO"] = "ERROR_NOMBRE"
                    errores_hoja += 1
                    continue
            df.at[idx, "nombre"] = nombre_limpio

            # Apellidos
            df.at[idx, "apellido1"] = limpiar_nombre_propio(fila.get("apellido1", ""))
            df.at[idx, "apellido2"] = limpiar_nombre_propio(fila.get("apellido2", ""))

            # Empresa y textos libres
            df.at[idx, "empresa"] = empresa_limpia
            for campo in ["tipo", "area", "cargo", "nota", "direccion"]:
                df.at[idx, campo] = limpiar_texto(fila.get(campo, ""))

            # Emails — hasta 3 campos (MAIL1, MAIL2, MAIL3)
            for campo_mail in ["MAIL1", "MAIL2", "MAIL3"]:
                email_limpio, email_valido, _ = limpiar_email(fila.get(campo_mail, ""))
                if email_valido is False:
                    avisos_fila.append(f"{campo_mail} inválido ('{fila.get(campo_mail,'')}') → se omite")
                    df.at[idx, campo_mail] = ""
                else:
                    df.at[idx, campo_mail] = email_limpio

            # Teléfonos — hasta 3 campos (TLF1, TLF2, TLF3)
            for campo_tlf in ["TLF1", "TLF2", "TLF3"]:
                tlf_limpio, tlf_valido, tlf_msg = limpiar_telefono(fila.get(campo_tlf, ""))
                df.at[idx, campo_tlf] = tlf_limpio
                if tlf_valido is False:
                    avisos_fila.append(f"{campo_tlf} inválido ('{fila.get(campo_tlf,'')}') → se omite ese teléfono")
                    df.at[idx, campo_tlf] = ""

            # Web
            df.at[idx, "web"] = limpiar_web(fila.get("web", ""))

            # Año — convertir columna a object para admitir strings y None
            df["año_ultimo_trabajo"] = df["año_ultimo_trabajo"].astype(object)
            anio_limpio, anio_valido, anio_msg = limpiar_anio(fila.get("año_ultimo_trabajo", ""))
            df.at[idx, "año_ultimo_trabajo"] = anio_limpio
            if anio_valido is False:
                avisos_fila.append(f"Año inválido: '{fila.get('año_ultimo_trabajo','')}' → se omite")
                df.at[idx, "año_ultimo_trabajo"] = None

            # Acciones — texto libre, solo limpiar
            for ac in ["accion1","accion2","accion3","accion4","accion5","accion6"]:
                df.at[idx, ac] = limpiar_texto(fila.get(ac, ""))

            # País (solo hoja H)
            if "pais" in df.columns:
                df.at[idx, "pais"] = limpiar_texto(fila.get("pais", ""))

            # Registrar avisos de esta fila
            if avisos_fila:
                avisos_hoja += 1
                log_lineas.append(f"  ⚠️  FILA {fila_num} ({nombre_limpio}):")
                for av in avisos_fila:
                    log_lineas.append(f"      → {av}")

            # Registrar emails en índice global para detección de duplicados
            for campo_mail in ["MAIL1", "MAIL2", "MAIL3"]:
                em = str(df.at[idx, campo_mail]).strip().lower()
                if em and em != "nan":
                    if em not in emails_globales:
                        emails_globales[em] = []
                    emails_globales[em].append({
                        "hoja": nombre_interno,
                        "fila": fila_num,
                        "nombre": nombre_limpio,
                        "campo": campo_mail,
                        "idx": idx
                    })

        log_lineas.append(f"  ✅ Procesada: {len(df)} filas | ❌ Errores: {errores_hoja} | ⚠️  Avisos: {avisos_hoja}\n")
        hojas_procesadas[nombre_interno] = df

    # ── Detección de duplicados entre hojas ────────────────────
    log_lineas.append(f"{'='*60}")
    log_lineas.append("DETECCIÓN DE DUPLICADOS POR EMAIL")
    log_lineas.append(f"{'='*60}")

    duplicados_encontrados = 0
    for email, apariciones in emails_globales.items():
        if len(apariciones) > 1:
            duplicados_encontrados += 1
            hojas_implicadas = [a["hoja"] for a in apariciones]
            categorias_combinadas = " | ".join(hojas_implicadas)
            log_lineas.append(
                f"  ℹ️  AVISO DUPLICADO: '{apariciones[0]['nombre']}' ({email})\n"
                f"      Aparece en: {', '.join(hojas_implicadas)}\n"
                f"      Acción: se mantiene en ambas hojas con categorías: {categorias_combinadas}\n"
                f"      → Revisa si hay datos distintos entre copias y decide cuál es la correcta."
            )

    if duplicados_encontrados == 0:
        log_lineas.append("  ✅ No se encontraron duplicados entre hojas.\n")
    else:
        log_lineas.append(f"\n  Total duplicados detectados: {duplicados_encontrados}\n")

    # ── Guardar Excel limpio ────────────────────────────────────
    log_lineas.append(f"{'='*60}")
    log_lineas.append("GUARDANDO EXCEL LIMPIO")
    log_lineas.append(f"{'='*60}")

    os.makedirs(os.path.dirname(RUTA_EXCEL_LIMPIO), exist_ok=True) if os.path.dirname(RUTA_EXCEL_LIMPIO) else None

    # Orden de columnas deseado en el Excel limpio
    ORDEN_COLUMNAS = [
        "nombre", "apellido1", "apellido2", "empresa", "tipo", "area", "cargo",
        "MAIL1", "MAIL2", "MAIL3", "TLF1", "TLF2", "TLF3",
        "año_ultimo_trabajo", "nota", "direccion", "web",
        "accion1", "accion2", "accion3", "accion4", "accion5", "accion6",
        "pais", "ACTIVO", "categoria"
    ]

    with pd.ExcelWriter(RUTA_EXCEL_LIMPIO, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas_procesadas.items():
            # Excluir filas con errores críticos
            df_valido = df[~df["ACTIVO"].str.startswith("ERROR", na=False)].copy()
            # Eliminar columna interna de número de fila (no va al Excel de salida)
            df_valido = df_valido.drop(columns=["__fila_excel__"], errors="ignore")
            # Reordenar columnas: primero las del orden deseado que existan, luego el resto
            cols_ordenadas = [c for c in ORDEN_COLUMNAS if c in df_valido.columns]
            cols_resto = [c for c in df_valido.columns if c not in ORDEN_COLUMNAS]
            df_valido = df_valido[cols_ordenadas + cols_resto]
            df_valido.to_excel(writer, sheet_name=nombre_hoja[:31], index=False)
            log_lineas.append(f"  ✅ {nombre_hoja}: {len(df_valido)} contactos guardados")

    log_lineas.append(f"\n✅ Excel limpio guardado en: {RUTA_EXCEL_LIMPIO}")

    # ── Guardar log ─────────────────────────────────────────────
    os.makedirs(os.path.dirname(RUTA_LOG), exist_ok=True) if os.path.dirname(RUTA_LOG) else None
    with open(RUTA_LOG, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lineas))

    print(f"\n✅ Proceso completado.")
    print(f"   Excel limpio: {RUTA_EXCEL_LIMPIO}")
    print(f"   Log:          {RUTA_LOG}")

    # Mostrar resumen en pantalla
    print("\n" + "\n".join([l for l in log_lineas if any(s in l for s in ["✅","❌","⚠️","ℹ️","HOJA"])]))


if __name__ == "__main__":
    procesar_excel()
